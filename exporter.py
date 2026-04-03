"""
Roblox Group - Export "Experienced" Members to Excel
=====================================================
Usage:
    pip install requests openpyxl
    python roblox_experienced_members.py

Config:
    Set GROUP_ID and ROLE_NAME below before running.
"""

import requests
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIGURATION — edit these before running
# ─────────────────────────────────────────────
GROUP_ID  = 2568175        # <-- Replace with your Roblox group ID
ROLE_NAME = "Experienced Participant" # <-- Exact role name to filter (case-insensitive)
OUTPUT_FILE = "experienced_members.xlsx"
# ─────────────────────────────────────────────

BASE_URL = "https://groups.roblox.com/v1"


def get_group_roles(group_id: int) -> list[dict]:
    """Return all roles in the group."""
    url = f"{BASE_URL}/groups/{group_id}/roles"
    r = requests.get(url)
    r.raise_for_status()
    return r.json().get("roles", [])


def find_role(roles: list[dict], name: str) -> dict | None:
    """Find a role by name (case-insensitive)."""
    for role in roles:
        if role["name"].strip().lower() == name.strip().lower():
            return role
    return None


def get_members_in_role(group_id: int, role_id: int) -> list[dict]:
    """Paginate through all members with the given role."""
    members = []
    cursor = ""
    page = 1

    while True:
        url = f"{BASE_URL}/groups/{group_id}/roles/{role_id}/users"
        params = {"limit": 100, "sortOrder": "Asc"}
        if cursor:
            params["cursor"] = cursor

        print(f"  Fetching page {page}...", end=" ", flush=True)
        r = requests.get(url, params=params)
        r.raise_for_status()
        data = r.json()

        batch = data.get("data", [])
        members.extend(batch)
        print(f"{len(batch)} members fetched (total: {len(members)})")

        cursor = data.get("nextPageCursor")
        if not cursor:
            break

        page += 1
        time.sleep(0.3)  # polite rate limiting

    return members


def build_spreadsheet(members: list[dict], group_id: int, role_name: str, output_file: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Experienced Members"

    # ── Styles ──────────────────────────────────────────────────────────────
    DARK_BLUE  = "1F3864"
    MID_BLUE   = "2E75B6"
    LIGHT_BLUE = "D6E4F0"
    WHITE      = "FFFFFF"
    ALT_ROW    = "EBF3FB"

    header_font  = Font(name="Arial", bold=True, color=WHITE, size=11)
    header_fill  = PatternFill("solid", start_color=DARK_BLUE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font  = Font(name="Arial", bold=True, color=WHITE, size=14)
    title_fill  = PatternFill("solid", start_color=MID_BLUE)
    title_align = Alignment(horizontal="center", vertical="center")

    meta_font  = Font(name="Arial", italic=True, color="444444", size=10)
    meta_align = Alignment(horizontal="left", vertical="center")

    data_font  = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def alt_fill(row_idx):
        color = ALT_ROW if row_idx % 2 == 0 else WHITE
        return PatternFill("solid", start_color=color)

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Roblox Group {group_id} — {role_name} Members"
    ws["A1"].font  = title_font
    ws["A1"].fill  = title_fill
    ws["A1"].alignment = title_align
    ws.row_dimensions[1].height = 30

    # ── Meta row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   "
        f"Total members: {len(members)}"
    )
    ws["A2"].font = meta_font
    ws["A2"].alignment = meta_align
    ws["A2"].fill = PatternFill("solid", start_color=LIGHT_BLUE)
    ws.row_dimensions[2].height = 20

    # ── Column headers ───────────────────────────────────────────────────────
    headers = ["#", "User ID", "Username", "Display Name", "Profile URL"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[3].height = 22

    # ── Data rows ────────────────────────────────────────────────────────────
    for i, member in enumerate(members, start=1):
        row = i + 3
        user_id  = member.get("userId") or member.get("id", "")
        username = member.get("username", "")
        display  = member.get("displayName", username)
        profile  = f"https://www.roblox.com/users/{user_id}/profile" if user_id else ""

        values = [i, user_id, username, display, profile]
        fill = alt_fill(i)

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font   = data_font
            cell.fill   = fill
            cell.border = border
            cell.alignment = center_align if col <= 2 else data_align

        ws.row_dimensions[row].height = 18

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [6, 14, 24, 24, 46]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Freeze panes below header ─────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A3:E{len(members) + 3}"

    wb.save(output_file)
    print(f"\n✅  Saved → {output_file}")


def main():
    print(f"Fetching roles for group {GROUP_ID}...")
    roles = get_group_roles(GROUP_ID)
    print(f"  Found {len(roles)} roles: {[r['name'] for r in roles]}")

    role = find_role(roles, ROLE_NAME)
    if not role:
        print(f"\n❌  Role '{ROLE_NAME}' not found in group {GROUP_ID}.")
        print("    Available roles:", [r["name"] for r in roles])
        return

    print(f"\nRole found: '{role['name']}' (ID: {role['id']})")
    print(f"Fetching members with role '{role['name']}'...")

    members = get_members_in_role(GROUP_ID, role["id"])
    print(f"\nTotal '{role['name']}' members: {len(members)}")

    if not members:
        print("No members found — nothing to export.")
        return

    print(f"\nBuilding spreadsheet...")
    build_spreadsheet(members, GROUP_ID, ROLE_NAME, OUTPUT_FILE)


if __name__ == "__main__":
    main()